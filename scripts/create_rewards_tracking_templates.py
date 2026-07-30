from pathlib import Path
import csv
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

ROOT = Path('/opt/data/midwest-suppliers-site')
OUT_DIR = ROOT / 'customer-tracking'
OUT_DIR.mkdir(exist_ok=True)
XLSX = OUT_DIR / 'midwest-suppliers-customer-tracker.xlsx'
SIGNUP_PDF = OUT_DIR / 'midwest-rewards-signup-sheet.pdf'
SIGNUP_HTML = OUT_DIR / 'midwest-rewards-signup-sheet.html'
LOGO = ROOT / 'assets' / 'knife-logo.jpg'
PHONE = '605-675-9429'
WEBSITE = 'justicejuly.github.io/midwest-suppliers-site'

sheets = {
    'Customers': [
        'Customer ID','Date Added','First Name','Last Name','Phone','Email','Address','City / Area','ZIP','Customer Type','Preferred Contact','Product Interest','Last Order Date','Last Order Amount','Lifetime Spend','Rewards Status','Referral Code','Notes'
    ],
    'Orders': [
        'Order ID','Order Date','Customer ID','Customer Name','Phone','Order Type','Items Ordered','Order Amount','Payment Method','Delivery Area','Delivery Date','Delivery Status','Rewards Points Earned','Notes'
    ],
    'Follow Ups': [
        'Follow Up Date','Customer ID','Customer Name','Phone','Reason','Priority','Status','Last Contacted','Next Step','Notes'
    ],
    'Rewards': [
        'Customer ID','Customer Name','Phone','Rewards Points','Lifetime Points','Rewards Tier','Last Reward Earned','Last Reward Redeemed','Notes'
    ],
    'Referrals': [
        'Referral Date','Referring Customer ID','Referring Customer Name','New Customer Name','New Customer Phone','Referral Status','New Customer Order Amount','Reward Given','Notes'
    ],
    'Restaurant Leads': [
        'Business Name','Contact Name','Phone','Email','Address','City','Website','Lead Source','Products to Pitch','Priority','Status','Follow Up Date','Notes'
    ],
    'Lists': ['List Name','Value']
}

list_values = {
    'Customer Type': ['Home','Business','Restaurant','Lodge','Bar','Office','Other'],
    'Preferred Contact': ['Call','Text','Email'],
    'Product Interest': ['Beef','Pork','Chicken','Seafood','Combo','Freezer','Business Order','Restaurant Weekly Order','Custom'],
    'Payment Method': ['Card','Cash App','Venmo','Check','Cash','Other'],
    'Delivery Status': ['Pending','Scheduled','Delivered','Cancelled'],
    'Follow Up Status': ['Not Started','Contacted','Done','No Answer','Do Not Contact'],
    'Follow Up Priority': ['High','Medium','Low'],
    'Rewards Status': ['New','Active','VIP','Inactive'],
    'Rewards Tier': ['Starter','Stocked','VIP'],
    'Referral Status': ['New','Contacted','Ordered','Rewarded','Dead'],
    'Restaurant Lead Status': ['New','Contacted','Meeting Set','Sample Dropped','Won','Lost'],
    'Restaurant Lead Priority': ['A','B','C'],
    'Order Type': ['Beef','Pork','Seafood','Chicken','Combo','Custom','Business'],
    'Follow Up Reason': ['New Lead','Reorder','Delivery Check-in','Restaurant Meeting','Referral','Reward','Quote Follow-up']
}

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill('solid', fgColor='B5122A')
header_font = Font(color='FFFFFF', bold=True)
sub_fill = PatternFill('solid', fgColor='F7F2E9')
thin = Side(style='thin', color='D8D0C5')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for name, headers in sheets.items():
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    for col_idx, header in enumerate(headers, 1):
        width = max(12, min(28, len(header) + 4))
        if header in ['Notes','Items Ordered','Next Step','Products to Pitch']:
            width = 34
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in range(2, 52):
        for col in range(1, len(headers)+1):
            ws.cell(row, col).border = border
            ws.cell(row, col).alignment = Alignment(vertical='top', wrap_text=True)

# Lists tab content and named-ish ranges via explicit ranges
lists_ws = wb['Lists']
row = 2
ranges = {}
for list_name, values in list_values.items():
    start = row
    for value in values:
        lists_ws.cell(row, 1).value = list_name
        lists_ws.cell(row, 2).value = value
        lists_ws.cell(row, 1).border = border
        lists_ws.cell(row, 2).border = border
        row += 1
    end = row - 1
    ranges[list_name] = f"'Lists'!$B${start}:$B${end}"

# Data validations helper
def add_dropdown(ws, header_name, list_name, max_row=500):
    headers = [c.value for c in ws[1]]
    if header_name not in headers or list_name not in ranges:
        return
    col = headers.index(header_name) + 1
    letter = get_column_letter(col)
    dv = DataValidation(type='list', formula1=f'={ranges[list_name]}', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'{letter}2:{letter}{max_row}')

# Apply dropdowns
add_dropdown(wb['Customers'], 'Customer Type', 'Customer Type')
add_dropdown(wb['Customers'], 'Preferred Contact', 'Preferred Contact')
add_dropdown(wb['Customers'], 'Product Interest', 'Product Interest')
add_dropdown(wb['Customers'], 'Rewards Status', 'Rewards Status')
add_dropdown(wb['Orders'], 'Order Type', 'Order Type')
add_dropdown(wb['Orders'], 'Payment Method', 'Payment Method')
add_dropdown(wb['Orders'], 'Delivery Status', 'Delivery Status')
add_dropdown(wb['Follow Ups'], 'Reason', 'Follow Up Reason')
add_dropdown(wb['Follow Ups'], 'Priority', 'Follow Up Priority')
add_dropdown(wb['Follow Ups'], 'Status', 'Follow Up Status')
add_dropdown(wb['Rewards'], 'Rewards Tier', 'Rewards Tier')
add_dropdown(wb['Referrals'], 'Referral Status', 'Referral Status')
add_dropdown(wb['Restaurant Leads'], 'Priority', 'Restaurant Lead Priority')
add_dropdown(wb['Restaurant Leads'], 'Status', 'Restaurant Lead Status')

# Sample row + formulas: write directly into row 2 so examples are visible immediately.
def write_row(ws, row_number, values):
    for col_number, value in enumerate(values, 1):
        cell = ws.cell(row_number, col_number)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

write_row(wb['Customers'], 2, ['CUST-0001', date.today().isoformat(), 'Test', 'Customer', '605-555-0000', '', '123 Sample St', 'Rapid City', '57701', 'Home', 'Text', 'Combo', '', '', '', 'New', 'TEST0000', 'Sample row — replace or delete'])
write_row(wb['Orders'], 2, ['ORD-0001', date.today().isoformat(), 'CUST-0001', 'Test Customer', '605-555-0000', 'Combo', 'Preferred Customer Combo', 999, 'Card', 'Rapid City', (date.today()+timedelta(days=1)).isoformat(), 'Scheduled', '=FLOOR(H2/25,1)', 'Sample row — delete before use'])
write_row(wb['Follow Ups'], 2, [(date.today()+timedelta(days=1)).isoformat(), 'CUST-0001', 'Test Customer', '605-555-0000', 'Delivery Check-in', 'Medium', 'Not Started', '', 'Check satisfaction and ask for referral', 'Sample follow-up'])
write_row(wb['Rewards'], 2, ['CUST-0001', 'Test Customer', '605-555-0000', '=SUMIF(Orders!C:C,A2,Orders!M:M)', '=D2', 'Starter', '', '', 'Sample rewards account'])
write_row(wb['Referrals'], 2, [date.today().isoformat(), 'CUST-0001', 'Test Customer', 'New Friend', '605-555-1111', 'New', '', '', 'Sample referral'])
write_row(wb['Restaurant Leads'], 2, ['Sample Steakhouse', 'Chef/GM', '605-555-2222', '', 'Main St', 'Rapid City', '', 'Prospecting', 'Steak, burger, seafood', 'A', 'New', (date.today()+timedelta(days=3)).isoformat(), 'Sample restaurant lead'])

# Re-apply styles to all data-entry rows.
for ws in wb.worksheets:
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

# Add instructions sheet first
instructions = wb.create_sheet('Start Here', 0)
instructions['A1'] = 'Midwest Suppliers Customer Tracker'
instructions['A1'].font = Font(size=18, bold=True, color='B5122A')
instructions['A3'] = 'Rewards Program Rules'
instructions['A3'].font = Font(bold=True)
instructions['A4'] = 'Earn 1 point for every $25 spent.'
instructions['A5'] = '5 points = $10 off. 10 points = $25 off. Referral: after first order, both customers get $25 off.'
instructions['A7'] = 'How to use this file'
instructions['A7'].font = Font(bold=True)
instructions['A8'] = '1. Add every new lead/customer to Customers.'
instructions['A9'] = '2. Add every purchase to Orders.'
instructions['A10'] = '3. Put every callback/reminder in Follow Ups.'
instructions['A11'] = '4. Track points in Rewards and referrals in Referrals.'
instructions['A12'] = '5. Keep restaurants/businesses in Restaurant Leads until they become customers.'
instructions['A14'] = f'Business phone: {PHONE}'
instructions['A15'] = f'Website: {WEBSITE}'
instructions.column_dimensions['A'].width = 100

wb.save(XLSX)

# CSV backups
for name, headers in sheets.items():
    csv_path = OUT_DIR / f"{name.lower().replace(' ', '-').replace('/', '-')}.csv"
    with csv_path.open('w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        if name == 'Lists':
            for list_name, vals in list_values.items():
                for val in vals:
                    writer.writerow([list_name, val])

# Printable signup sheet HTML
SIGNUP_HTML.write_text(f'''<!doctype html>
<html><head><meta charset="utf-8"><title>Midwest Rewards Signup Sheet</title>
<style>
@page{{size:letter;margin:.35in}}body{{font-family:Arial,Helvetica,sans-serif;color:#17110d;margin:0}}.page{{display:grid;grid-template-columns:1fr 1fr;gap:.22in}}.card{{border:1px solid #17110d;padding:.18in;min-height:4.85in;position:relative}}h1{{font-size:19px;margin:0;text-transform:uppercase}}h2{{font-size:13px;margin:6px 0;color:#b5122a;text-transform:uppercase}}p{{font-size:10px;margin:5px 0;color:#5d5147}}.phone{{font-size:17px;color:#b5122a;font-weight:bold}}.line{{border-bottom:1px solid #333;height:.28in;margin:6px 0 9px}}.choices{{font-size:10px;line-height:1.55}}.footer{{position:absolute;bottom:.12in;left:.18in;right:.18in;border-top:1px solid #d8d0c5;padding-top:4px;font-size:9px;color:#5d5147}}@media print{{body{{margin:0}}}}
</style></head><body><div class="page">
''' + '\n'.join([f'''<section class="card"><h1>Midwest Rewards Club</h1><p>Earn rewards when you order. Refer a friend — after their first order, you both get $25 off.</p><div class="phone">{PHONE}</div><h2>Customer signup</h2>Name:<div class="line"></div>Phone:<div class="line"></div>Email:<div class="line"></div>Delivery Area:<div class="line"></div><div class="choices"><b>Interested in:</b> Beef / Pork / Chicken / Seafood / Combo / Business<br><b>Preferred contact:</b> Call / Text / Email<br><b>Referred by:</b> ______________________________</div>Notes:<div class="line" style="height:.55in"></div><div class="footer">Midwest Suppliers Meats & Seafood · {WEBSITE}</div></section>''' for _ in range(4)]) + '</div></body></html>')

# Signup PDF
styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name='SignupTitle', fontName='Helvetica-Bold', fontSize=16, leading=17, textColor=colors.HexColor('#17110d')))
styles.add(ParagraphStyle(name='SignupHead', fontName='Helvetica-Bold', fontSize=11, leading=12, textColor=colors.HexColor('#b5122a')))
styles.add(ParagraphStyle(name='SignupSmall', fontName='Helvetica', fontSize=8.5, leading=10, textColor=colors.HexColor('#5d5147')))
styles.add(ParagraphStyle(name='SignupPhone', fontName='Helvetica-Bold', fontSize=15, leading=16, textColor=colors.HexColor('#b5122a')))

def P(txt, style='SignupSmall'):
    return Paragraph(txt, styles[style])

def signup_card():
    lines = []
    lines.append(P('Midwest Rewards Club','SignupTitle'))
    lines.append(P('Earn rewards when you order. Refer a friend — after their first order, you both get $25 off.'))
    lines.append(P(PHONE,'SignupPhone'))
    lines.append(P('Customer signup','SignupHead'))
    line_style = TableStyle([('LINEBELOW',(0,0),(-1,-1),0.7,colors.black),('BOTTOMPADDING',(0,0),(-1,-1),10),('TOPPADDING',(0,0),(-1,-1),6)])
    for label in ['Name:', 'Phone:', 'Email:', 'Delivery Area:']:
        t=Table([[P(label)]], colWidths=[3.15*inch]); t.setStyle(line_style); lines.append(t)
    lines.append(P('<b>Interested in:</b> Beef / Pork / Chicken / Seafood / Combo / Business'))
    lines.append(P('<b>Preferred contact:</b> Call / Text / Email'))
    t=Table([[P('Referred by:')]], colWidths=[3.15*inch]); t.setStyle(line_style); lines.append(t)
    t=Table([[P('Notes:')]], colWidths=[3.15*inch], rowHeights=[.45*inch]); t.setStyle(line_style); lines.append(t)
    lines.append(P(f'Midwest Suppliers Meats & Seafood · {WEBSITE}'))
    outer = Table([[lines]], colWidths=[3.45*inch], rowHeights=[4.85*inch])
    outer.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.8,colors.HexColor('#17110d')),('LEFTPADDING',(0,0),(-1,-1),10),('RIGHTPADDING',(0,0),(-1,-1),10),('TOPPADDING',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),8),('VALIGN',(0,0),(-1,-1),'TOP')]))
    return outer

pdf = SimpleDocTemplate(str(SIGNUP_PDF), pagesize=letter, leftMargin=.35*inch, rightMargin=.35*inch, topMargin=.35*inch, bottomMargin=.35*inch, title='Midwest Rewards Signup Sheet')
pdf.build([Table([[signup_card(), signup_card()], [signup_card(), signup_card()]], colWidths=[3.8*inch,3.8*inch], rowHeights=[5.0*inch,5.0*inch], style=[('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)])])

print(XLSX)
print(SIGNUP_PDF)
print(SIGNUP_HTML)
print(OUT_DIR)
